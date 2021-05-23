import openpyxl

# def write_data(file,sheetname,row_name,col_num,data):
#     wb=openpyxl.load_workbook("C://asha//colaco//asha.xlsx")
#     sheet = wb[sheetname]
#     sheet.cell(row=row_name, column=col_num).value=data
#     wb.save("C://asha//colaco//asha.xlsx")
# write_data("asha.xlsx","Sheet1",3,2,"agnelo")


# def read_data():  # how to read data from specific coloumn and multiple rows
#     data={}
#     wb=openpyxl.load_workbook("C://asha//colaco//asha.xlsx")
#     sheet=wb["Sheet1"]
#     for i in range(1,sheet.max_row+1): #we are taking the the range of row as it is not fixed whereas coloumn is fixed
#         data[sheet.cell(row=i,column=1).value]=sheet.cell(row=i,column=2).value
#     print(data)
# z=read_data()

def read_data():  # how to read data from entire file
    data=[]
    wb=openpyxl.load_workbook("C://asha//colaco//asha.xlsx")
    sheet=wb["Sheet1"]
    for i in range(1, sheet.max_row + 1):
        for j in range(1,sheet.max_column+1):
            num=sheet.cell(row=i,column=j).value
            if num is not None:
                data.append(num)
    print(data)
output=read_data()


# #print max row and print specific value
# def read_data(file,sheetname):
#     wb=openpyxl.load_workbook("C://asha//colaco//asha.xlsx")
#     sheet=wb[sheetname]
#     new=sheet.max_row
#     print(new)
#     val=sheet["A1"].value
#     print(val)
# z=read_data("asha.xlsx","Sheet1")



